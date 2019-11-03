"""
Script to move addresses from excel to label
template in word.
"""
from __future__ import print_function
from mailmerge import MailMerge
import openpyxl

# load workbook
wb = openpyxl.load_workbook('guest list(1).xlsx')
list = wb['new list']
row_count = list.max_row + 1
row_start = 2

# get access to word doc with the template/table
# using https://pbpython.com/python-word-template.html
template = "test.docx"
doc = MailMerge(template)

# store dictionary in list
label_data = []

for i in range (row_start, row_count):
	name = list.cell(row=i, column=1).value

	# remove leading and trailing whitespace, capitalize
	street = list.cell(row=i, column=2).value.strip().upper()
	city = list.cell(row=i, column=3).value.strip().upper()
	state = list.cell(row=i, column=4).value.strip().upper()
	zip = list.cell(row=i, column=5).value.strip()

	line3 = "%s, %s %s" % (city, state, zip)

	label_dict = {
		'name' : name,
		'address' : street,
		'city' : line3
		}
	label_data.append(label_dict)

doc.merge_rows('name', label_data)
doc.write('test1.docx')