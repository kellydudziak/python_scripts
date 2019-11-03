"""
Python script to parse guest list from excel.
"""

import openpyxl
import re					# regex

# load workbook and name sheets
wb = openpyxl.load_workbook('guest list.xlsx')
old_list = wb['guest list']
new_list = wb['new list']

row_count = old_list.max_row + 1
col_count = 3					# only name and address from old_list
row_start = 2					# start at 2 to ignore the headers

# which columns hold which component
name_col = 1
street_col = 2
city_col = 3
state_col = 4
zip_col = 5

for i in range(row_start, row_count):
	# copy name over
	nm = old_list.cell(row=i, column=1).value
	name = re.sub('and', '&\n', nm)
	new_list.cell(row=i, column=1, value=name)

	full_addr = old_list.cell(row=i, column=2).value

	# pull state
	st = re.findall(r" [a-zA-Z][a-zA-Z] | (?i)florida", full_addr)
	if not st:					# if list is empty
		state = 'FL'
	elif len(st) > 1:				# if more than one result is returned, we want the last one
		state = st[len(st)-1]
	else:						# contains exactly one element
		state = st[0]
	new_list.cell(row=i, column=state_col, value=state)


	# pull zip
	zip = re.findall(r"\d{5}", full_addr)
	if not zip:
		zipcode = '00000'
	#elif len(zip) > 1:
		#zipcode = zip[len(zip)-1]
	else:
		zipcode = zip[len(zip)-1]
	new_list.cell(row=i, column=zip_col, value=zipcode)
	
	
	# pull street address and city
	split_addr = full_addr.split(',')

	street = split_addr[0]
	new_list.cell(row=i, column=street_col, value=street)

	city = split_addr[1]
	new_list.cell(row=i, column=city_col, value=city)	

wb.save('guest list(1).xlsx')